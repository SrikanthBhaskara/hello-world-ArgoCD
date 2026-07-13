import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_file_info(directory):
    """Scan a directory and collect file metadata."""
    records = []
    for root, dirs, files in os.walk(directory):
        dirs[:] = [d for d in dirs if d.lower() not in ("trash", ".trash", "$recycle.bin", ".recycle")]
        for file_name in files:
            full_path = os.path.abspath(os.path.join(root, file_name))
            _, ext = os.path.splitext(file_name)
            file_type = ext.lstrip(".").upper() if ext else "UNKNOWN"
            try:
                size_bytes = os.path.getsize(full_path)
                if size_bytes < 1024:
                    file_size = f"{size_bytes} B"
                elif size_bytes < 1024 ** 2:
                    file_size = f"{size_bytes / 1024:.2f} KB"
                else:
                    file_size = f"{size_bytes / (1024 ** 2):.2f} MB"
            except OSError:
                file_size = "N/A"
            records.append({
                "File_Name": file_name,
                "File_Type": file_type,
                "File_Size": file_size,
                "Full_File_Path": full_path,
                "Remarks": "",
            })
    return records


def export_to_excel(records, output_file="file_inventory.xlsx"):
    """Write file records to a formatted Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "File Inventory"

    headers = ["S.No", "File_Name", "File_Type", "File_Size", "Full_File_Path", "Remarks"]

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Alternate row colours
    light_blue = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Write data rows
    for row_idx, record in enumerate(records, start=2):
        row_fill = light_blue if row_idx % 2 == 0 else white_fill
        values = [
            row_idx - 1,
            record["File_Name"],
            record["File_Type"],
            record["File_Size"],
            record["Full_File_Path"],
            record["Remarks"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = center_align if col_idx in (1, 3, 4) else left_align
            cell.font = Font(name="Calibri", size=10)

    # Auto-fit column widths
    col_widths = [6, 35, 12, 12, 70, 25]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_file)
    print(f"Excel file saved: {os.path.abspath(output_file)}")


if __name__ == "__main__":
    # Scan the directory where this script is located
    scan_directory = os.path.dirname(os.path.abspath(__file__))
    output_excel = os.path.join(scan_directory, "personal_file_inventory_sree.xlsx")

    print(f"Scanning: {scan_directory}")
    records = get_file_info(scan_directory)
    print(f"Found {len(records)} file(s).")
    export_to_excel(records, output_excel)
